from __future__ import print_function
import random
import numpy as np
import os
import torch
from collections import deque
from collections import defaultdict, deque
from game import Board, Game
from mcts_pure import MCTSPlayer as MCTS_Pure
from mcts_alphaZero import MCTSPlayer
from mcts_alphaZero_BTMM_update_evaluation import MCTSPlayerBTMM
from policy_value_net_pytorch import PolicyValueNet 
import time
import pandas as pd

class ExplanationEvaluator:

    def create_test_data(self):
        """Tạo dữ liệu test để kiểm tra việc lưu file"""
        print("[DEBUG] Creating test evaluation data...")
        
        test_data = [
            {
                "game_id": 1, "move_id": 1, "game_phase": "opening",
                "state_hash": 12345, "move_count": 1,
                "explanation_size": 8, "explanation_time": 0.002,
                "action_agreement": 1, "surrogate_fidelity": 0.8,
                "comprehensiveness": 0.6, "sufficiency": 0.9,
                "chosen_action": 15, "policy_top_action": 15, "btmm_top_action": 15
            },
            {
                "game_id": 1, "move_id": 2, "game_phase": "opening", 
                "state_hash": 12346, "move_count": 2,
                "explanation_size": 10, "explanation_time": 0.003,
                "action_agreement": 0, "surrogate_fidelity": 0.6,
                "comprehensiveness": 0.4, "sufficiency": 0.7,
                "chosen_action": 22, "policy_top_action": 20, "btmm_top_action": 22
            }
        ]
        
        self.evaluation_data.extend(test_data)
        print(f"[DEBUG] Added {len(test_data)} test rows, total: {len(self.evaluation_data)}")
    
  
    def __init__(self, board_width, board_height, n_in_row):
        self.board_width = board_width
        self.board_height = board_height
        self.n_in_row = n_in_row
        self.evaluation_data = []
        self.explanation_times = []
        
    def classify_game_phase(self, board, move_count):
        """
        Phân loại phase của game dựa trên số nước đi
        """
        total_positions = board.width * board.height
        filled_positions = bin(board.bitboards[1] | board.bitboards[2]).count("1")
        fill_ratio = filled_positions / total_positions
        
        if move_count <= 6:  # 3 moves mỗi người
            return "opening"
        elif fill_ratio < 0.7:
            return "midgame"
        else:
            return "endgame"

    def compute_comprehensiveness_sufficiency(self, chosen_action, btmm_probs, policy_probs):
        """
        Tính Comprehensiveness và Sufficiency của explanation
        
        Args:
            chosen_action: hành động được chọn
            btmm_probs: dict {action: probability} từ BTMM explanation
            policy_probs: list [(action, probability)] từ policy network
            
        Returns:
            comprehensiveness, sufficiency
        """
        # Tìm xác suất gốc của action được chọn
        original_prob = next((prob for action, prob in policy_probs if action == chosen_action), 0.0)
        
        # Lấy top-k features từ explanation (các actions quan trọng nhất)
        k = min(5, len(btmm_probs))
        top_k_actions = sorted(btmm_probs.items(), key=lambda x: x[1], reverse=True)[:k]
        top_k_set = set(action for action, prob in top_k_actions)
        
        # COMPREHENSIVENESS: Loại bỏ top-k features và đo độ giảm xác suất
        remaining_probs = [prob for action, prob in policy_probs if action not in top_k_set]
        if remaining_probs:
            total_remaining = sum(remaining_probs)
            if total_remaining > 0:
                comprehensiveness_prob = next((prob/total_remaining for action, prob in policy_probs 
                                             if action == chosen_action and action not in top_k_set), 0.0)
            else:
                comprehensiveness_prob = 0.0
        else:
            comprehensiveness_prob = 0.0
            
        comprehensiveness = max(0, original_prob - comprehensiveness_prob)
        
        # SUFFICIENCY: Chỉ giữ lại top-k features và đo khả năng khôi phục
        sufficiency_probs = [prob for action, prob in policy_probs if action in top_k_set]
        if sufficiency_probs:
            total_sufficiency = sum(sufficiency_probs)
            if total_sufficiency > 0:
                sufficiency_prob = next((prob/total_sufficiency for action, prob in policy_probs 
                                       if action == chosen_action and action in top_k_set), 0.0)
            else:
                sufficiency_prob = 0.0
        else:
            sufficiency_prob = 0.0
            
        sufficiency = sufficiency_prob
        
        return comprehensiveness, sufficiency

    def compute_explanation_metrics(self, state, chosen_action, btmm_probs, policy_probs, move_count):
        """
        Tính toán các metrics đánh giá lời giải thích
        """
        start_time = time.time()
        
        # 1. Compactness - Average Explanation Size (AES)
        explanation_size = len(btmm_probs)
        
        # 2. Tính explanation time
        explanation_time = time.time() - start_time
        self.explanation_times.append(explanation_time)
        
        # 3. Action Agreement - So sánh policy network và BTMM
        policy_top_action = max(policy_probs, key=lambda x: x[1])[0] if policy_probs else chosen_action
        btmm_top_action = max(btmm_probs.items(), key=lambda x: x[1])[0] if btmm_probs else chosen_action
        action_agreement = 1 if policy_top_action == btmm_top_action else 0
        
        # 4. Surrogate Fidelity
        agreement_count = 0
        total_comparisons = min(5, len(btmm_probs))
        
        btmm_sorted = sorted(btmm_probs.items(), key=lambda x: x[1], reverse=True)[:total_comparisons]
        policy_sorted = sorted(policy_probs, key=lambda x: x[1], reverse=True)[:total_comparisons]
        
        btmm_actions = [action for action, _ in btmm_sorted]
        policy_actions = [action for action, _ in policy_sorted]
        
        for action in btmm_actions:
            if action in policy_actions:
                agreement_count += 1
                
        surrogate_fidelity = agreement_count / total_comparisons if total_comparisons > 0 else 0
        
        # Comprehensiveness và Sufficiency
        comprehensiveness, sufficiency = self.compute_comprehensiveness_sufficiency(
            chosen_action, btmm_probs, policy_probs
        )
        
        metrics = {
            "explanation_size": explanation_size,
            "explanation_time": explanation_time,
            "action_agreement": action_agreement,
            "surrogate_fidelity": surrogate_fidelity,
            "comprehensiveness": comprehensiveness,
            "sufficiency": sufficiency,
            "chosen_action": chosen_action,
            "policy_top_action": policy_top_action,
            "btmm_top_action": btmm_top_action
        }
        
        print(f"[METRICS] Comp: {comprehensiveness:.3f}, Suff: {sufficiency:.3f}, Agreement: {action_agreement}, Fidelity: {surrogate_fidelity:.3f}")
        
        return metrics

    def log_evaluation(self, game_id, move_id, state, metrics, move_count):
        """Ghi log kết quả đánh giá"""
        game_phase = self.classify_game_phase(state, move_count)
        
        evaluation_entry = {
            "game_id": game_id,
            "move_id": move_id,
            "game_phase": game_phase,
            "state_hash": hash(str(state.current_state())),
            "move_count": move_count,
            **metrics
        }
        self.evaluation_data.append(evaluation_entry)

    def calculate_summary_metrics(self):
        """Tính toán summary metrics theo group"""
        if not self.evaluation_data:
            return None
            
        df = pd.DataFrame(self.evaluation_data)
        
        # Group by game phase
        groups = {
            'All': df,
            'Opening': df[df['game_phase'] == 'opening'],
            'Midgame': df[df['game_phase'] == 'midgame'], 
            'Endgame': df[df['game_phase'] == 'endgame']
        }
        
        summary_rows = []
        for group_name, group_df in groups.items():
            if len(group_df) == 0:
                continue
                
            summary = {
                'Group': group_name,
                'Count': len(group_df),
                'Mean_Explanation_Size': group_df['explanation_size'].mean(),
                'Median_Explanation_Size': group_df['explanation_size'].median(),
                'Std_Explanation_Size': group_df['explanation_size'].std(),
                'Mean_Explanation_Time': group_df['explanation_time'].mean(),
                'Std_Explanation_Time': group_df['explanation_time'].std(),
                'Mean_Action_Agreement': group_df['action_agreement'].mean(),
                'Mean_Surrogate_Fidelity': group_df['surrogate_fidelity'].mean(),
                'Mean_Comprehensiveness': group_df['comprehensiveness'].mean(),
                'Mean_Sufficiency': group_df['sufficiency'].mean()
            }
            summary_rows.append(summary)
        
        return pd.DataFrame(summary_rows)

    
        
    def save_detailed_evaluation(self, filename="evaluation_results_detailed.xlsx", save_plots=True):
        """
        Lưu kết quả evaluation với cấu trúc giống maze:
        - PerState_Metrics (từng lượt/move)
        - PerState_WithGroups (thêm cột group based on game_phase: opening/midgame/endgame)
        - Group_Summary (thống kê theo group)
        - Summary (overall aggregates)
        - Surrogate_Info (thông tin surrogate nếu có)
        """
        import pandas as pd
        import numpy as np
        import os
        import matplotlib.pyplot as plt
        from openpyxl import load_workbook

        print(f"[DEBUG] save_detailed_evaluation called with filename: {filename}")
        print(f"[DEBUG] evaluation_data length: {len(self.evaluation_data)}")

        if not self.evaluation_data:
            print("[DEBUG] Không có dữ liệu evaluation để lưu")
            return None

        # 1) DataFrame cơ bản
        df = pd.DataFrame(self.evaluation_data).copy()
        # Chuẩn hoá tên cột nếu cần
        # Một số cột quan trọng mong đợi: explanation_size, explanation_time, action_agreement, surrogate_fidelity, chosen_action, policy_top_action, btmm_top_action, game_phase, move_count, game_id, move_id, state_hash
        expected_cols = ['game_id','move_id','game_phase','state_hash','move_count',
                         'explanation_size','explanation_time','action_agreement','surrogate_fidelity',
                         'comprehensiveness','sufficiency',
                         'chosen_action','policy_top_action','btmm_top_action']
        # đảm bảo tồn tại các cột
        for c in expected_cols:
            if c not in df.columns:
                df[c] = np.nan

        # Sanitize object columns to strings (list/dict/ndarray) to avoid Excel serializing issues
        for col in df.columns:
            # if any element is list/dict/ndarray, convert whole column to str
            try:
                sample = df[col].dropna().iloc[0]
                if isinstance(sample, (list, dict, np.ndarray)):
                    df[col] = df[col].apply(lambda x: str(x) if not pd.isna(x) else x)
            except Exception:
                # empty column or other issue -> skip
                pass

        # 2) PerState_Metrics sheet -> ghi trực tiếp df (các cột raw)
        per_state_df = df.copy()

        # 3) Grouping theo game_phase: Opening/Midgame/Endgame
        # Chuẩn hoá giá trị game_phase (lower)
        per_state_df['game_phase'] = per_state_df['game_phase'].astype(str).str.lower()
        # map to canonical groups
        def _phase_map(x):
            if 'open' in x:
                return 'Opening'
            elif 'mid' in x:
                return 'Midgame'
            elif 'end' in x:
                return 'Endgame'
            else:
                return 'Unknown'
        per_state_df['Group'] = per_state_df['game_phase'].apply(_phase_map)

        # 4) Tính thêm metric AES (ở Connect6 AES = explanation_size; MSX analog)
        per_state_df['AES'] = per_state_df['explanation_size'].astype(float)

        # 5) Thêm vài cột phụ tiện lợi
        # Action agreement (cột đã có), but also add if chosen==policy_top, chosen==btmm_top
        per_state_df['Chosen_eq_PolicyTop'] = (per_state_df['chosen_action'] == per_state_df['policy_top_action']).astype(int)
        per_state_df['Chosen_eq_BtmmTop'] = (per_state_df['chosen_action'] == per_state_df['btmm_top_action']).astype(int)

        # 6) PerState_WithGroups: copy per_state_df but ensure certain columns present
        perstate_with_groups = per_state_df.copy()

        # 7) Group_Summary: for groups All, Opening, Midgame, Endgame, Unknown
        groups = {
            'All': perstate_with_groups.index >= 0,
            'Opening': perstate_with_groups['Group'] == 'Opening',
            'Midgame': perstate_with_groups['Group'] == 'Midgame',
            'Endgame': perstate_with_groups['Group'] == 'Endgame',
            'Unknown': perstate_with_groups['Group'] == 'Unknown'
        }

        metric_cols = ['AES', 'explanation_time', 'action_agreement', 'surrogate_fidelity', 
                      'comprehensiveness', 'sufficiency', 'Chosen_eq_PolicyTop', 'Chosen_eq_BtmmTop']
        summary_rows = []
        for gname, mask in groups.items():
            sel = perstate_with_groups[mask]
            summary = {'Group': gname, 'Count': int(len(sel))}
            for mc in metric_cols:
                if mc in sel.columns:
                    vals = sel[mc].dropna().astype(float)
                    summary[f'{mc}_mean'] = float(vals.mean()) if len(vals) > 0 else float('nan')
                    summary[f'{mc}_median'] = float(vals.median()) if len(vals) > 0 else float('nan')
                    summary[f'{mc}_std'] = float(vals.std()) if len(vals) > 0 else float('nan')
                else:
                    summary[f'{mc}_mean'] = float('nan')
                    summary[f'{mc}_median'] = float('nan')
                    summary[f'{mc}_std'] = float('nan')
            summary_rows.append(summary)
        df_group_summary = pd.DataFrame(summary_rows)

        # 8) Overall Summary sheet
        overall = {
            'Total_Rows': len(perstate_with_groups),
            'Mean_AES': float(perstate_with_groups['AES'].dropna().mean()) if 'AES' in perstate_with_groups.columns and len(perstate_with_groups['AES'].dropna())>0 else float('nan'),
            'Median_AES': float(perstate_with_groups['AES'].dropna().median()) if 'AES' in perstate_with_groups.columns and len(perstate_with_groups['AES'].dropna())>0 else float('nan'),
            'Mean_Explanation_Time': float(perstate_with_groups['explanation_time'].dropna().mean()) if 'explanation_time' in perstate_with_groups.columns and len(perstate_with_groups['explanation_time'].dropna())>0 else float('nan'),
            'Mean_Action_Agreement': float(perstate_with_groups['action_agreement'].dropna().mean()) if 'action_agreement' in perstate_with_groups.columns and len(perstate_with_groups['action_agreement'].dropna())>0 else float('nan'),
            'Mean_Surrogate_Fidelity': float(perstate_with_groups['surrogate_fidelity'].dropna().mean()) if 'surrogate_fidelity' in perstate_with_groups.columns and len(perstate_with_groups['surrogate_fidelity'].dropna())>0 else float('nan'),
            'Mean_Comprehensiveness': float(perstate_with_groups['comprehensiveness'].dropna().mean()) if 'comprehensiveness' in perstate_with_groups.columns and len(perstate_with_groups['comprehensiveness'].dropna())>0 else float('nan'),
            'Mean_Sufficiency': float(perstate_with_groups['sufficiency'].dropna().mean()) if 'sufficiency' in perstate_with_groups.columns and len(perstate_with_groups['sufficiency'].dropna())>0 else float('nan'),
        }
        df_summary = pd.DataFrame([overall])

        # 9) Surrogate_Info (chỉ có thể báo những gì ta biết)
        surrogate_info = {
            'NumStatesEvaluated': len(perstate_with_groups),
            'Notes': 'Evaluation includes comprehensiveness (impact of removing important features) and sufficiency (ability to recover decision with only important features) metrics.'
        }
        df_surrogate_info = pd.DataFrame([surrogate_info])

        # 10) Lưu Excel (giữ cấu trúc giống maze)
        os.makedirs(os.path.dirname(filename) if os.path.dirname(filename) else '.', exist_ok=True)

        # If file exists, remove prior sheets we will overwrite to avoid "sheet exists" errors.
        try:
            if os.path.exists(filename):
                wb = load_workbook(filename)
                for sheet in ['PerState_Metrics', 'PerState_WithGroups', 'Group_Summary', 'Summary', 'Surrogate_Info']:
                    if sheet in wb.sheetnames:
                        wb.remove(wb[sheet])
                wb.save(filename)
        except Exception as e:
            # Non-fatal: print warn and continue (we'll write with mode='a' or 'w')
            print(f"[WARN] Could not pre-clean existing workbook: {e}")

        try:
            # append to file (file exists or not) - now target sheets removed
            with pd.ExcelWriter(filename, engine='openpyxl', mode='a' if os.path.exists(filename) else 'w') as writer:
                per_state_df.to_excel(writer, sheet_name='PerState_Metrics', index=False)
                perstate_with_groups.to_excel(writer, sheet_name='PerState_WithGroups', index=False)
                df_group_summary.to_excel(writer, sheet_name='Group_Summary', index=False)
                df_summary.to_excel(writer, sheet_name='Summary', index=False)
                df_surrogate_info.to_excel(writer, sheet_name='Surrogate_Info', index=False)
        except Exception as e:
            print(f"[ERROR] Writing Excel failed: {e}")
            # As fallback, try writing fresh file (overwrite)
            try:
                with pd.ExcelWriter(filename, engine='openpyxl', mode='w') as writer:
                    per_state_df.to_excel(writer, sheet_name='PerState_Metrics', index=False)
                    perstate_with_groups.to_excel(writer, sheet_name='PerState_WithGroups', index=False)
                    df_group_summary.to_excel(writer, sheet_name='Group_Summary', index=False)
                    df_summary.to_excel(writer, sheet_name='Summary', index=False)
                    df_surrogate_info.to_excel(writer, sheet_name='Surrogate_Info', index=False)
            except Exception as e2:
                print(f"[CRITICAL] Fallback write also failed: {e2}")
                return None

        print(f"[✓] Detailed evaluation saved to {filename}")

        # 11) VẼ vài biểu đồ tương tự maze: AES hist per group, boxplot explanation_time, bar group mean AES
        try:
            outdir = os.path.dirname(filename) or '.'
            # AES hist per group - FIXED to avoid scalar/max issues
            plt.figure(figsize=(8,4))
            for gname, mask in groups.items():
                sel = perstate_with_groups[mask]
                aes_vals = sel['AES'].dropna() if 'AES' in sel.columns else pd.Series(dtype=float)
                if len(aes_vals) > 0:
                    max_bin = int(np.nanmax(aes_vals)) if not np.isnan(np.nanmax(aes_vals)) else 0
                    bins = range(0, max_bin + 2)
                    plt.hist(aes_vals, bins=bins, alpha=0.4, label=gname)
            plt.legend()
            plt.title('AES distribution by group')
            plt.tight_layout()
            plt.savefig(os.path.join(outdir, 'group_aes_hist_connect6.png'))
            plt.close()
        except Exception as e:
            print(f"[WARN] Could not plot AES hist: {e}")

        try:
            # Bar: mean AES per group
            plt.figure(figsize=(6,4))
            x = df_group_summary['Group']
            y = df_group_summary['AES_mean'] if 'AES_mean' in df_group_summary.columns else None
            if y is not None:
                plt.bar(x, y.fillna(0))
                plt.title('Mean AES by Group')
                plt.tight_layout()
                plt.savefig(os.path.join(outdir, 'group_mean_aes_connect6.png'))
                plt.close()
        except Exception as e:
            print(f"[WARN] Could not plot group mean AES: {e}")

        try:
            # Scatter plot Comprehensiveness vs Sufficiency
            plt.figure(figsize=(8,6))
            plt.scatter(perstate_with_groups['comprehensiveness'], 
                       perstate_with_groups['sufficiency'], 
                       alpha=0.6, c='blue')
            plt.xlabel('Comprehensiveness')
            plt.ylabel('Sufficiency')
            plt.title('Comprehensiveness vs Sufficiency')
            plt.grid(True, alpha=0.3)
            plt.tight_layout()
            plt.savefig(os.path.join(outdir, 'comp_vs_suff_scatter.png'))
            plt.close()
        except Exception as e:
            print(f"[WARN] Could not plot comp vs suff scatter: {e}")

        return df_summary, df_group_summary, perstate_with_groups


class TrainPipelineWithEvaluation():
    def __init__(self, init_model=None, board_width=6, board_height=6,
                 n_in_row=4, n_playout=400, use_gpu=False, is_shown=False,
                 output_file_name="", game_batch_number=1500,
                 use_rws=False, pure_use_rws=False, time_limit=None,
                 enable_evaluation=True):
        
        self.time_limit = time_limit 
        self.board_width = board_width
        self.board_height = board_height
        self.n_in_row = n_in_row
        self.board = Board(width=self.board_width,
                           height=self.board_height,
                           n_in_row=self.n_in_row)
        self.game = Game(self.board)
        self.learn_rate = 2e-3 
        self.lr_multiplier = 1.0
        self.temp = 1.0
        self.n_playout = n_playout
        self.c_puct = 5
        self.buffer_size = 10000
        self.batch_size = 512
        self.data_buffer = deque(maxlen=self.buffer_size)
        self.play_batch_size = 1
        self.epochs = 5
        self.kl_targ = 0.02
        self.check_freq = 1
        self.game_batch_num = game_batch_number
        self.best_win_ratio = 0.0
        self.pure_mcts_playout_num = 1000
        self.use_gpu = use_gpu
        self.is_shown = is_shown
        self.output_file_name = output_file_name
        self.use_rws = use_rws           
        self.pure_use_rws = pure_use_rws 
        self.enable_evaluation = enable_evaluation
        
        # Khởi tạo explanation evaluator
        if self.enable_evaluation:
            self.evaluator = ExplanationEvaluator(board_width, board_height, n_in_row)
        
        self.policy_value_net = PolicyValueNet(self.board_width,
                                               self.board_height,
                                               model_file=init_model,
                                               use_gpu=self.use_gpu
                                               )
        self.mcts_player = MCTSPlayerBTMM(self.policy_value_net.policy_value_fn,
                                        c_puct=self.c_puct,
                                        n_playout=self.n_playout,
                                        use_rws=True, time_limit=self.time_limit)
        self.game_id = None
        self.current_iter = 0 

    def get_equi_data(self, play_data):
        extend_data = []
        for state, mcts_porb, winner in play_data:
            for i in [1, 2, 3, 4]:
                equi_state = np.array([np.rot90(s, i) for s in state])
                equi_mcts_prob = np.rot90(np.flipud(
                    mcts_porb.reshape(self.board_height, self.board_width)), i)
                extend_data.append((equi_state,
                                    np.flipud(equi_mcts_prob).flatten(),
                                    winner))
                equi_state = np.array([np.fliplr(s) for s in equi_state])
                equi_mcts_prob = np.fliplr(equi_mcts_prob)
                extend_data.append((equi_state,
                                    np.flipud(equi_mcts_prob).flatten(),
                                    winner))
        return extend_data

    def collect_selfplay_data(self, n_games=1):
        """
        Thu thập dữ liệu tự chơi với evaluation explanations
        """
        for i in range(n_games):
            game_id = i + 1 
            board = Board(width=self.board_width, height=self.board_height, n_in_row=self.n_in_row)
            game = Game(board)
            mcts_player = MCTSPlayerBTMM(
                self.policy_value_net.policy_value_fn,
                c_puct=self.c_puct,
                n_playout=self.n_playout,
                use_rws=True,
                time_limit=self.time_limit
            )
            mcts_player.game_id = game_id 
            game.game_id = game_id
      
            if self.enable_evaluation:
                winner, play_data = self._collect_selfplay_with_evaluation(game, mcts_player, game_id)
            else:
                winner, play_data = game.start_self_play(mcts_player, temp=self.temp, game_id=game_id)

            play_data = list(play_data)
            self.episode_len = len(play_data)
            play_data = self.get_equi_data(play_data)
            self.data_buffer.extend(play_data)

    def _collect_selfplay_with_evaluation(self, game, mcts_player, game_id):
        """
        Phiên bản self-play có thu thập explanation metrics
        """
        board = game.board
        board.init_board()
        states, mcts_probs, current_players = [], [], []
        move_count = 0
        
        while True:
            move_count += 1
            
            # Lấy policy probabilities từ network
            policy_probs, _ = self.policy_value_net.policy_value_fn(board)
            
            # Lấy move từ MCTS BTMM
            move, move_probs = mcts_player.get_action(board,
                                                    temp=self.temp,
                                                    return_prob=True,
                                                    game_id=game_id,
                                                    move_id=move_count)
      
            if hasattr(mcts_player.mcts, '_root') and mcts_player.mcts._root._children:
                # Lấy BTMM probabilities từ children nodes
                btmm_probs = {}
                for action, node in mcts_player.mcts._root._children.items():
                    btmm_probs[action] = node._n_visits / max(1, mcts_player.mcts._root._n_visits)
                
                # Tính explanation metrics (ĐÃ CẬP NHẬT với comprehensiveness & sufficiency)
                metrics = self.evaluator.compute_explanation_metrics(
                    board, move, btmm_probs, policy_probs, move_count
                )
                
                # Log evaluation
                self.evaluator.log_evaluation(game_id, move_count, board, metrics, move_count)
            
            # Thực hiện move
            states.append(board.current_state())
            mcts_probs.append(move_probs)
            current_players.append(board.get_current_player())
            board.do_move(move)
            
            end, winner = board.game_end()
            if end:
                winners_z = np.zeros(len(current_players))
                if winner != -1:
                    winners_z[np.array(current_players) == winner] = 1.0
                    winners_z[np.array(current_players) != winner] = -1.0
                # Reset player
                mcts_player.reset_player()
                
                if winner != -1:
                    print(f"[Game {game_id}] Kết thúc. Người thắng: {winner}")
                else:
                    print(f"[Game {game_id}] Kết thúc. Hòa")
                
          
                play_data = zip(states, mcts_probs, winners_z)
                return winner, play_data

    

    def policy_update(self):
        mini_batch = random.sample(self.data_buffer, self.batch_size)
        state_batch = [data[0] for data in mini_batch]
        mcts_probs_batch = [data[1] for data in mini_batch]
        winner_batch = [data[2] for data in mini_batch]
        old_probs, old_v = self.policy_value_net.policy_value(state_batch)
        for i in range(self.epochs):
            loss, entropy = self.policy_value_net.train_step(
                    state_batch,
                    mcts_probs_batch,
                    winner_batch,
                    self.learn_rate*self.lr_multiplier)
            new_probs, new_v = self.policy_value_net.policy_value(state_batch)
            kl = np.mean(np.sum(old_probs * (
                    np.log(old_probs + 1e-10) - np.log(new_probs + 1e-10)),
                    axis=1)
            )
            if kl > self.kl_targ * 4:
                break
        if kl > self.kl_targ * 2 and self.lr_multiplier > 0.1:
            self.lr_multiplier /= 1.5
        elif kl < self.kl_targ / 2 and self.lr_multiplier < 10:
            self.lr_multiplier *= 1.5

        explained_var_old = (1 -
                             np.var(np.array(winner_batch) - old_v.flatten()) /
                             np.var(np.array(winner_batch)))
        explained_var_new = (1 -
                             np.var(np.array(winner_batch) - new_v.flatten()) /
                             np.var(np.array(winner_batch)))
        print(("kl:{:.5f},"
               "lr_multiplier:{:.3f},"
               "loss:{}," 
               "entropy:{}," 
               "explained_var_old:{:.3f},"
               "explained_var_new:{:.3f}"
               ).format(kl,
                        self.lr_multiplier,
                        loss,
                        entropy,
                        explained_var_old,
                        explained_var_new))
        return loss, entropy

    def policy_evaluate(self, n_games=10):
        current_mcts_player = MCTSPlayerBTMM(self.policy_value_net.policy_value_fn,
                                            c_puct=self.c_puct,
                                            n_playout=self.n_playout,
                                            use_rws=True,
                                            time_limit=self.time_limit)

        pure_mcts_player = MCTS_Pure(c_puct=5,
                                    n_playout=self.pure_mcts_playout_num,
                                    use_rws=self.pure_use_rws)
        win_cnt = defaultdict(int)
        for i in range(n_games):
            winner = self.game.start_play(current_mcts_player,
                                        pure_mcts_player,
                                        start_player=i % 2,
                                        is_shown=self.is_shown)
            win_cnt[winner] += 1
            print(f"[Ván {i+1}] Kết quả: {winner}")
        win_ratio = 1.0 * (win_cnt[1] + 0.5 * win_cnt[-1]) / n_games
        print("Số lượt mô phỏng:{}, Thắng: {}, Thua: {}, Hòa:{}".format(
                self.pure_mcts_playout_num,
                win_cnt[1], win_cnt[2], win_cnt[-1]))
        return win_ratio

    def run(self):
        os.makedirs("checkpoints", exist_ok=True)
        os.makedirs("info", exist_ok=True)
        os.makedirs("evaluation", exist_ok=True)

        # THÊM: Tạo và lưu test data NGAY từ đầu
        if self.enable_evaluation and hasattr(self, 'evaluator'):
            print("[DEBUG] Creating and saving test data immediately...")
            self.evaluator.create_test_data()
            
            # LƯU FILE NGAY LẬP TỨC
            test_filename = f"evaluation/test_data_{self.output_file_name}.xlsx"
            res = self.evaluator.save_detailed_evaluation(test_filename)
            
            if res is not None:
                df_summary, df_group_summary, perstate_with_groups = res
                print(f"[SUCCESS] Test file saved: {test_filename}")
                try:
                    print("[DEBUG] Summary:")
                    print(df_summary.to_string(index=False))
                except Exception:
                    pass
            else:
                print("[ERROR] Failed to save test file")

        loss_log_path = "info/" + str(self.board) + "_loss_" + self.output_file_name + ".txt"
        win_ratio_log_path = "info/" + str(self.board) + "_win_ration" + self.output_file_name + ".txt"
        
        if self.current_iter == 0:
            with open(loss_log_path, 'w', encoding="utf-8") as loss_file:
                loss_file.write("Số lần tự chơi,loss,entropy\n")
            with open(win_ratio_log_path, 'w', encoding="utf-8") as win_ratio_file:
                win_ratio_file.write("Số lần tự chơi, Số lượt mô phỏng pure_MCTS, Tỷ lệ thắng\n")

        try:
            for i in range(self.current_iter, self.game_batch_num):
                self.collect_selfplay_data(self.play_batch_size)
                print("Ván tự chơi i:{}, Số bước đã đi:{}".format(i+1, self.episode_len))

                if len(self.data_buffer) > self.batch_size:
                    loss, entropy = self.policy_update()
                    with open(loss_log_path, 'a', encoding='utf-8') as loss_file:
                        loss_file.write(f"{i+1},{loss},{entropy}\n")

                if (i+1) % self.check_freq == 0:
                    print("Số ván tự chơi hiện tại: {}".format(i+1))
                    win_ratio = self.policy_evaluate()
                    with open(win_ratio_log_path, 'a', encoding='utf-8') as win_ratio_file:
                        win_ratio_file.write(f"{i+1},{self.pure_mcts_playout_num},{win_ratio}\n")

                    # Lưu evaluation results
                    if self.enable_evaluation and hasattr(self, 'evaluator') and self.evaluator.evaluation_data:
                        eval_filename = f"evaluation/eval_{self.output_file_name}_iter_{i+1}.xlsx"
                        res = self.evaluator.save_detailed_evaluation(eval_filename)
                        
                        if res is not None:
                            df_summary, df_group_summary, perstate_with_groups = res
                            print(f"[EVALUATION] Evaluation summary tại iter {i+1}:")
                            try:
                                print(df_summary.to_string(index=False))
                            except Exception:
                                pass

                    self.policy_value_net.save_model('./model/' + str(self.board_height)
                                                    + '_' + str(self.board_width)
                                                    + '_' + str(self.n_in_row) +
                                                    '_current_policy_' + self.output_file_name + '.model')
                    if win_ratio >= self.best_win_ratio:
                        print("Đã tạo ra chiến lược tốt hơn!!!!!!!!")
                        self.best_win_ratio = win_ratio
                        self.policy_value_net.save_model('./model/' + str(self.board_height)
                                                        + '_' + str(self.board_width)
                                                        + '_' + str(self.n_in_row) +
                                                        '_best_policy_' + self.output_file_name + '.model')
                        if (self.best_win_ratio == 1.0 and self.pure_mcts_playout_num < 50000):
                            self.pure_mcts_playout_num += 1000
                            self.best_win_ratio = 0.0
                self.current_iter = i + 1
                if (i + 1) % 1 == 0:
                    self.save_checkpoint(f"checkpoints/session_{i+1}.pth")
      

        finally:
 
            if self.enable_evaluation and hasattr(self, 'evaluator'):
                print(f"[DEBUG] Final evaluation data count: {len(self.evaluator.evaluation_data)}")
                
                if self.evaluator.evaluation_data:
                    # Tạo file CSV đơn giản
                    try:
                        import csv
                        csv_filename = f"evaluation/final_evaluation_{self.output_file_name}.csv"
                        os.makedirs("evaluation", exist_ok=True)
                        
                        with open(csv_filename, 'w', newline='', encoding='utf-8') as csvfile:
                            fieldnames = self.evaluator.evaluation_data[0].keys()
                            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                            writer.writeheader()
                            writer.writerows(self.evaluator.evaluation_data)
                        
                        print(f"[SUCCESS] ĐÃ LƯU FILE CSV: {csv_filename}")
                        print(f"[SUCCESS] Số dòng dữ liệu: {len(self.evaluator.evaluation_data)}")
                        
                        # In thử vài dòng đầu để kiểm tra
                        print("\n[DEBUG] Sample data (first 3 rows):")
                        for i, row in enumerate(self.evaluator.evaluation_data[:3]):
                            print(f"Row {i}: {row}")
                            
                    except Exception as e:
                        print(f"[ERROR] Không thể lưu CSV: {e}")
                else:
                    print("[DEBUG] Không có evaluation data để lưu")

        
            # Lưu evaluation data cuối cùng
            if self.enable_evaluation and hasattr(self, 'evaluator'):
                self.evaluator.save_detailed_evaluation(f"evaluation/final_evaluation_{self.output_file_name}.xlsx")

            if self.enable_evaluation and hasattr(self, 'evaluator'):
                self.evaluator.create_test_data()


    def save_checkpoint(self, filename):
        checkpoint = {
            'model_state_dict': self.policy_value_net.get_policy_param(),
            'optimizer_state_dict': self.policy_value_net.optimizer.state_dict(),
            'data_buffer': list(self.data_buffer),
            'current_iter': self.current_iter,
            'best_win_ratio': self.best_win_ratio,
        }
        if self.enable_evaluation:
            checkpoint['evaluation_data'] = self.evaluator.evaluation_data
            checkpoint['explanation_times'] = self.evaluator.explanation_times
            
        torch.save(checkpoint, filename)
        print(f"[INFO] Đã lưu checkpoint: {filename}")

    def load_checkpoint(self, filename):
        checkpoint = torch.load(filename, map_location='cpu')
        self.policy_value_net.policy_value_net.load_state_dict(checkpoint['model_state_dict'])
        self.policy_value_net.optimizer.load_state_dict(checkpoint['optimizer_state_dict'])
        self.data_buffer = deque(checkpoint['data_buffer'], maxlen=self.buffer_size)
        self.current_iter = checkpoint.get('current_iter', 0)
        self.best_win_ratio = checkpoint.get('best_win_ratio', 0.0)
        
        if self.enable_evaluation and 'evaluation_data' in checkpoint:
            self.evaluator.evaluation_data = checkpoint['evaluation_data']
            self.evaluator.explanation_times = checkpoint['explanation_times']
            
        print(f"[INFO] Đã load checkpoint: {filename}, tiếp tục từ ván {self.current_iter + 1}")


def usage():
    print("-s Thiết lập kích thước bàn cờ, mặc định là 6")
    print("-r Thiết lập số quân liên tiếp để thắng, mặc định là 4")
    print("-m Thiết lập số lượt mô phỏng MCTS mỗi nước đi, mặc định là 400")
    print("-o Định danh file lưu model đã huấn luyện (lưu ý: chương trình sẽ tự động tạo tên file dựa trên tham số model)")
    print("-n Thiết lập số ván tự chơi để huấn luyện, mặc định là 1500")
    print("--use_gpu Sử dụng GPU để huấn luyện")
    print("--graphics Hiển thị giao diện khi đánh giá mô hình")
    print("--rws      Bật chế độ RWS cho agent AlphaZero")
    print("--pure_rws Bật chế độ RWS cho agent Pure MCTS")
    print("--enable_eval Bật đánh giá explanations (mặc định: True)")
    print("--disable_eval Tắt đánh giá explanations")

if __name__ == '__main__':
    import sys, getopt
    import os

    height = 10
    width = 10
    n_in_row = 6
    use_gpu = False
    n_playout = 800
    is_shown = False
    output_file_name = ""
    game_batch_number = 1500
    init_model_name = None
    battle = False
    use_rws = False        
    pure_use_rws = False   
    resume = False
    resume_checkpoint_file = None
    time_limit = None
    enable_evaluation = True

    opts, args = getopt.getopt(
    sys.argv[1:], "hs:r:m:go:n:i:",
    ["use_gpu", "graphics", "rws", "pure_rws", "btmm", "time_limit=", "enable_eval", "disable_eval"]
    )

    for op, value in opts:
        if op == "-h":
            usage()
            sys.exit()
        elif op == "-s":
            height = width = int(value)
        elif op == "-r":
            n_in_row = int(value)
        elif op == "--use_gpu":
            use_gpu = True
        elif op == "-m":
            n_playout = int(value)
        elif op == "-g" or op == "--graphics":
            is_shown = True
        elif op == "-o":
            output_file_name = value
        elif op == "-i":
            init_model_name = value
        elif op == "-n":
            game_batch_number = int(value)
        elif op == "--rws":          
            use_rws = True
        elif op == "--pure_rws":     
            pure_use_rws = True
        elif op == "--resume":
            resume = True
            resume_checkpoint_file = value
        elif op == "--time_limit":
            time_limit = float(value)
        elif op == "--enable_eval":   
            enable_evaluation = True
        elif op == "--disable_eval":  
            enable_evaluation = False

    training_pipeline = TrainPipelineWithEvaluation(
        board_height=height, board_width=width,
        n_in_row=n_in_row, use_gpu=use_gpu,
        n_playout=n_playout, is_shown=is_shown,
        output_file_name=output_file_name,
        init_model=init_model_name,
        game_batch_number=game_batch_number,
        use_rws=use_rws,
        pure_use_rws=pure_use_rws,
        time_limit=time_limit,
        enable_evaluation=enable_evaluation
    )
    if resume and resume_checkpoint_file and os.path.exists(resume_checkpoint_file):
        training_pipeline.load_checkpoint(resume_checkpoint_file)
    training_pipeline.run()