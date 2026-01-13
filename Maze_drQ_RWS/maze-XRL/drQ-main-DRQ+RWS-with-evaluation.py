import numpy as np
import random
import matplotlib.pyplot as plt
import sys
import pandas as pd 
import os  
from collections import deque
import matplotlib.pyplot as plt
import numpy as np
import time
import ast
import math
from collections import Counter



def generate_maze(grid_size, filename, density=0.25):
    while True:
        maze = []
        for i in range(grid_size):
            row = []
            for j in range(grid_size):
                if (i, j) == (0, 0) or (i, j) == (grid_size - 1, grid_size - 1):
                    row.append('.')
                else:
                    row.append('#' if random.random() < density else '.')
            maze.append(row)

        # Open diagonal path from (0,0) to (n-1,n-1)
        for i in range(grid_size):
            maze[i][i] = '.'

        # Open additional horizontal path in the middle
        mid = grid_size // 2
        for j in range(grid_size):
            maze[mid][j] = '.'

        # Convert to blocked_points list
        blocked = [(i, j) for i in range(grid_size) for j in range(grid_size) if maze[i][j] == '#']

        # Check if path exists
        if is_path_available((0, 0), (grid_size - 1, grid_size - 1), blocked, grid_size):
            break  # OK, path exists

    # Write to file
    with open(filename, 'w') as f:
        for row in maze:
            f.write(' '.join(row) + '\n')
    print(f"Maze {grid_size}x{grid_size} saved to {filename}")

class RewardPathFinder:
    def __init__(self, grid_size, maze_file):
        self.grid_size = grid_size
        self.start = (0, 0)
        self.end = (grid_size - 1, grid_size - 1)
        self.blocked_points = self.load_maze(maze_file)
        
        self.reward_components = ['turn', 'goal', 'blocked', 'safe']
        self.num_components = len(self.reward_components)
        
        self.q_table = np.zeros((self.num_components, grid_size, grid_size, 4))
        
        self.epsilon = 1.0
        self.alpha = 0.1
        self.gamma = 0.9
        self.best_path = []
        self.consecutive_safe_actions = 0
        self.reward_history = []
        self.steps_history = []
        self.final_reward = 0
        self.final_steps = 0
        self.final_turns = 0
        self.episodes = 0  # in __init__

            
    # Load maze matrix from file
    def load_maze(self, maze_file):
        blocked_points = []
        with open(maze_file, 'r') as file:
            for i, line in enumerate(file):
                for j, char in enumerate(line.strip().split()):
                    if char == '#':
                        blocked_points.append((i, j))
        return blocked_points

    def reset(self):
        return self.start

    def is_terminal(self, state):
        return state == self.end
    
    # Get all feasible actions from current state
    def get_actions(self, state):
        actions = []
        for action in range(4):
            next_state = self.take_action(state, action)
            if 0 <= next_state[0] < self.grid_size and 0 <= next_state[1] < self.grid_size:
                actions.append(action)
        return actions

    # Execute action and return next state
    def take_action(self, state, action):
        if action == 0:  # up
            return (state[0] - 1, state[1])
        elif action == 1:  # down
            return (state[0] + 1, state[1])
        elif action == 2:  # left
            return (state[0], state[1] - 1)
        elif action == 3:  # right
            return (state[0], state[1] + 1)


    # Action selection function 
    def choose_action(self, state):
        x, y = state
        valid_actions = self.get_actions(state)
        if not valid_actions:
            return random.randint(0, 3)  # fallback if stuck

        # Exploration: choose randomly with epsilon probability
        if random.uniform(0, 1) < self.epsilon:
            return random.choice(valid_actions)

        # Calculate total Q from reward components
        total_q = np.sum(self.q_table[:, x, y, :], axis=0)

        # Get Q-values for valid actions
        q_values = np.array([total_q[a] for a in valid_actions])

        # If all Q are negative, shift to ensure non-negative
        min_q = np.min(q_values)
        if min_q < 0:
            q_values = q_values - min_q

        # If sum of Q = 0 (avoid division by 0), fallback random
        f_sum = np.sum(q_values)
        if f_sum == 0:
            return random.choice(valid_actions)

        # Calculate probability distribution: Q_i / total Q
        probs = q_values / f_sum

        # Select action according to Roulette Wheel distribution
        chosen_index = np.random.choice(len(valid_actions), p=probs)
        return valid_actions[chosen_index] 


    def extract_features(self, state, action):
        features = []
        next_state = self.take_action(state, action)
        x, y = next_state

        # 1. Near wall
        if x in [0, self.grid_size - 1] or y in [0, self.grid_size - 1]:
            features.append("BorderLevel_1")
        elif x in [1, self.grid_size - 2] or y in [1, self.grid_size - 2]:
            features.append("BorderLevel_2")
        else:
            features.append("BorderLevel_3")

        # 2. Turn penalty
        if hasattr(self, 'last_action') and self.last_action is not None:
            if abs(action - self.last_action) == 1 or abs(action - self.last_action) == 3:
                features.append("TurnPenalty")

        # 3. Near goal
        if abs(x - self.end[0]) + abs(y - self.end[1]) <= 2:
            features.append("GoalNear_1")

        # 4. Blocked near next_state
        for a in range(4):
            future = self.take_action(next_state, a)
            if future in self.blocked_points:
                features.append("BlockedNear_1")
                break

        # 5. Safe streak
        if self.consecutive_safe_actions >= 2:
            features.append("SafeStreak_2")

        return features

    # drQ - Decomposed Reward Q Algorithm    
    def compute_decomposed_rewards(self, state, action, next_state, last_action):
        # Initialize rewards for each component
        rewards = {c: 0 for c in self.reward_components}
        if last_action is not None and (
            (action == 0 and last_action == 1) or (action == 1 and last_action == 0) or
            (action == 2 and last_action == 3) or (action == 3 and last_action == 2)):
            rewards['turn'] = -1 # Penalty for turning back
        if self.is_terminal(next_state):
            rewards['goal'] = 30 # Reward for reaching goal
        if next_state in self.blocked_points:
            rewards['blocked'] = -1 # Penalty for hitting blocked point
            self.consecutive_safe_actions = 0
        else:
            for a in range(4):
                future_state = self.take_action(next_state, a)
                if (0 <= future_state[0] < self.grid_size and 0 <= future_state[1] < self.grid_size and
                    future_state in self.blocked_points):
                    rewards['blocked'] = -0.5 # Penalty for future collision with blocked point
                    break
            self.consecutive_safe_actions += 1
            if self.consecutive_safe_actions == 2:
                rewards['safe'] = 2 # Reward for consecutive safe moves
                self.consecutive_safe_actions = 0
        return rewards

    # Update Q-Table
    def update_q_table(self, state, action, rewards, next_state):
        total_q_next = np.sum(self.q_table[:, next_state[0], next_state[1], :], axis=0)
        best_next_action = np.argmax(total_q_next)
        for c_idx, component in enumerate(self.reward_components):
            r_c = rewards[component]
            q_c_next = self.q_table[c_idx, next_state[0], next_state[1], best_next_action]
            td_target = r_c + self.gamma * q_c_next
            td_delta = td_target - self.q_table[c_idx, state[0], state[1], action]
            self.q_table[c_idx, state[0], state[1], action] += self.alpha * td_delta

    # RDX - Reward Difference Explanation Algorithm
    def reward_difference_explanation(self, state, chosen_action, alternative_action):
        # Get Q-value of chosen action
        q_chosen = self.q_table[:, state[0], state[1], chosen_action]
        
        # Get Q-value of alternative action
        q_alternative = self.q_table[:, state[0], state[1], alternative_action]
        q_diff = q_chosen - q_alternative
        
        # Save difference between two actions for each reward component
        explanation = {}
        for c_idx, component in enumerate(self.reward_components):
            explanation[component] = q_diff[c_idx]
        
        # Calculate total difference
        total_diff = np.sum(q_diff)
        return explanation, total_diff

    # MSX - Minimal Sufficient Explanation Algorithm
    def minimal_sufficient_explanation(self, state, chosen_action, alternative_action):
        # Calculate difference between actions
        explanation, total_diff = self.reward_difference_explanation(state, chosen_action, alternative_action)
        
        # d represents how much negative components reduce the value of chosen action
        d = sum(abs(diff) for comp, diff in explanation.items() if diff < 0)
        
        # List of reward components with positive differences, sorted descending
        positive_components = [(comp, diff) for comp, diff in explanation.items() if diff > 0]
        positive_components.sort(key=lambda x: x[1], reverse=True)
        
        # Minimal sufficient positive components
        msx_plus = []
        current_sum = 0
        for comp, diff in positive_components:
            current_sum += diff
            msx_plus.append(comp)
            if current_sum > d:
                break
        
        if msx_plus:
            msx_plus_diffs = [explanation[comp] for comp in msx_plus]
            # v represents how much positive components in msx_plus exceed the minimum value
            v = sum(msx_plus_diffs) - min(msx_plus_diffs)
        else:
            v = 0
        
        # Identify negative components
        negative_components = [(comp, diff) for comp, diff in explanation.items() if diff < 0]
        negative_components.sort(key=lambda x: abs(x[1]), reverse=True)
        
        # Minimal sufficient negative components
        msx_minus = []
        current_sum = 0
        for comp, diff in negative_components:
            current_sum += -diff
            msx_minus.append(comp)
            if current_sum > v:
                break
        
        # Create formula details
        formula_details = {
            'd': d,
            'v': v,
            'msx_plus_sum': sum(explanation[comp] for comp in msx_plus) if msx_plus else 0,
            'msx_minus_sum': sum(-explanation[comp] for comp in msx_minus) if msx_minus else 0
        }
        
        return msx_plus, msx_minus, explanation, formula_details
    
    # Get run count from file
    def get_run_count(self, grid_size):
        # File to store run count
        run_count_file = f"run_count_{grid_size}.txt"
        
        # If file doesn't exist, initialize run count to 0
        if not os.path.exists(run_count_file):
            with open(run_count_file, 'w') as f:
                f.write("0")
        
        # Read run count from file
        with open(run_count_file, 'r') as f:
            run_count = int(f.read().strip())
        
        run_count += 1
        
        # Write new run count to file
        with open(run_count_file, 'w') as f:
            f.write(str(run_count))
        
        return run_count
    
    # Create output file name for results
    def prepare_output_file(self, grid_size, run_count):
        output_dir = "excel-results"
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        output_file = os.path.join(output_dir, f"evaluation-report-final.xlsx")
        return output_file


    def train(self, episodes, log_random_episode=False):
        self.reward_history = []
        self.steps_history = []
        best_total_reward = -float('inf')
        for episode in range(episodes):
            steps = 0
            state = self.reset()
            total_reward = 0
            current_path = [state]
            last_action = None
            self.consecutive_safe_actions = 0
            
            # Initialize Q-Table for new episode
            while not self.is_terminal(state):
                steps = 0  # count steps
                action = self.choose_action(state)
                next_state = self.take_action(state, action)
                current_path.append(next_state)
                
                rewards = self.compute_decomposed_rewards(state, action, next_state, last_action)
                total_reward += sum(rewards.values())
                self.update_q_table(state, action, rewards, next_state)
                
                self.steps_history.append(len(current_path))
                state = next_state
                last_action = action
                
                steps += 1
            self.reward_history.append(total_reward)
            self.steps_history.append(steps)
            if total_reward > best_total_reward:
                best_total_reward = total_reward
                self.best_path = current_path
            if episode == episodes - 1:
                self.final_reward = total_reward
                self.final_steps = len(current_path) - 1
               
                # Count turns (180 degree)
                turn_count = 0
                for i in range(2, len(current_path)):
                    dx1 = current_path[i-1][0] - current_path[i-2][0]
                    dy1 = current_path[i-1][1] - current_path[i-2][1]
                    dx2 = current_path[i][0] - current_path[i-1][0]
                    dy2 = current_path[i][1] - current_path[i-1][1]
                    if (dx1 == -dx2 and dy1 == -dy2):
                        turn_count += 1
                self.final_turns = turn_count
                print(f" Training episodes recorded: {len(self.reward_history)}")
            # Export data to Excel in the last episode
            if episode == episodes - 1 and log_random_episode:
                print(f"\nLogging actions for the last episode (Episode {episode + 1}) to Excel file...")
                table_data = []

                # Create output_file EARLY
                output_file = self.prepare_output_file(self.grid_size, self.get_run_count(self.grid_size))

                for i in range(self.grid_size):
                    for j in range(self.grid_size):
                        state = (i, j)
                        total_q = np.sum(self.q_table[:, state[0], state[1], :], axis=0)
                        chosen_action = np.argmax(total_q)
                        next_state = self.take_action(state, chosen_action)

                        available_actions = self.get_actions(state)
                        if len(available_actions) > 1:
                            alternative_action = random.choice([a for a in available_actions if a != chosen_action])
                        else:
                            alternative_action = chosen_action

                        msx_plus, msx_minus, explanation, formula_details = self.minimal_sufficient_explanation(
                            state, chosen_action, alternative_action)
                        features = self.extract_features(state, chosen_action)
                        gamma_multiplier = 1.0  # Not using gamma_dict anymore


                        # Calculate roulette probabilities for each component
                        probs_dict = self.get_roulette_probabilities(state)

                        row = {
                            "State": str(state),
                            "Chosen Action": chosen_action,
                            "Alternative Action": alternative_action,
                            "Δ_turn": round(explanation['turn'], 2),
                            "Δ_goal": round(explanation['goal'], 2),
                            "Δ_blocked": round(explanation['blocked'], 2),
                            "Δ_safe": round(explanation['safe'], 2),
                            "Total Δ": round(sum(explanation.values()), 2),
                            "d": round(formula_details['d'], 2),
                            "MSX+": str(msx_plus),
                            "Sum MSX+": round(formula_details['msx_plus_sum'], 2),
                            "v": round(formula_details['v'], 2),
                            "MSX-": str(msx_minus),
                            "Sum MSX-": round(formula_details['msx_minus_sum'], 2),
                            "Next State": str(next_state)
                        }

                        for comp in self.reward_components + ['total']:
                            row[f"{comp}_P"] = f"{{'Up': {probs_dict[comp][0]:.4f}, 'Down': {probs_dict[comp][1]:.4f}, 'Left': {probs_dict[comp][2]:.4f}, 'Right': {probs_dict[comp][3]:.4f}}}"

                        # Add features
                        table_data.append(row)
                
                # DRAW PLOTS (Reward + Heatmap)
                self.plot_metrics()
                self.visualize(show_q_values=True)

                # SAVE EXCEL FILE
                df = pd.DataFrame(table_data)
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Episode Log', index=False)

                    for c_idx, component in enumerate(self.reward_components):
                        q_table_component = self.q_table[c_idx]
                        q_data = []
                        for i in range(self.grid_size):
                            for j in range(self.grid_size):
                                row = {
                                    "State": str((i, j)),
                                    "Q_Up": round(q_table_component[i, j, 0], 2),
                                    "Q_Down": round(q_table_component[i, j, 1], 2),
                                    "Q_Left": round(q_table_component[i, j, 2], 2),
                                    "Q_Right": round(q_table_component[i, j, 3], 2)
                                }
                                q_data.append(row)
                        q_df = pd.DataFrame(q_data)
                        q_df.to_excel(writer, sheet_name=f'Q_{component}', index=False)

                    total_q_table = np.sum(self.q_table, axis=0)
                    total_q_data = []
                    for i in range(self.grid_size):
                        for j in range(self.grid_size):
                            row = {
                                "State": str((i, j)),
                                "Q_Up": round(total_q_table[i, j, 0], 2),
                                "Q_Down": round(total_q_table[i, j, 1], 2),
                                "Q_Left": round(total_q_table[i, j, 2], 2),
                                "Q_Right": round(total_q_table[i, j, 3], 2)
                            }
                            total_q_data.append(row)
                    total_q_df = pd.DataFrame(total_q_data)
                    total_q_df.to_excel(writer, sheet_name='Q_Total', index=False)

                print(f"Data has been exported to {output_file}")
                # Save roulette probabilities for each state
                roulette_rows = []
                for i in range(self.grid_size):
                    for j in range(self.grid_size):
                        state = (i, j)
                        probs = self.get_roulette_probabilities(state)

                        row = {"State": str(state)}
                        for comp in self.reward_components + ['total']:
                            for a_name, a_idx in zip(['Up', 'Down', 'Left', 'Right'], range(4)):
                                row[f"{comp}_P_{a_name}"] = round(probs[comp][a_idx], 4)

                        roulette_rows.append(row)
                        roulette_df = pd.DataFrame(roulette_rows)
                        roulette_df.to_excel(writer, sheet_name="Roulette_Probs", index=False)

            
            print(f"Episode {episode + 1}: Total Reward: {total_reward}")
            if episode == episodes - 1:
                print("Final Decomposed Q-Table:")
                for c_idx, component in enumerate(self.reward_components):
                    print(f"Component: {component}")
                    for i in range(self.grid_size):
                        for j in range(self.grid_size):
                            print(f"State {i, j}: Q-Values: {self.q_table[c_idx, i, j]}")
            
    # Get shortest path from start to end
    from collections import deque
    def get_shortest_path(self):
        from heapq import heappush, heappop

        visited = set()
        came_from = {}
        start = self.start
        end = self.end

        queue = []
        heappush(queue, (-np.max(np.sum(self.q_table[:, start[0], start[1], :], axis=0)), start))  # use maximum Q-value

        while queue:
            _, current = heappop(queue)
            if current == end:
                break

            visited.add(current)
            total_q = np.sum(self.q_table[:, current[0], current[1], :], axis=0)
            for action in range(4):
                next_state = self.take_action(current, action)
                if (0 <= next_state[0] < self.grid_size and
                    0 <= next_state[1] < self.grid_size and
                    next_state not in self.blocked_points and
                    next_state not in visited):
                    came_from[next_state] = current
                    heappush(queue, (-total_q[action], next_state))

        # reconstruct path
        path = []
        current = end
        while current != start:
            if current not in came_from:
                return []  # Path not found
            path.append(current)
            current = came_from[current]
        path.append(start)
        return path[::-1]  # reverse path

    # Draw heatmap for actions
    def visualize(self, show_q_values=False):
        total_q = np.sum(self.q_table, axis=0)
        q_values = np.max(total_q, axis=2)
        plt.imshow(q_values, cmap='hot', interpolation='nearest')
        if show_q_values:
            plt.colorbar(label='Total Q-Values')
        else:
            plt.colorbar(label='Heatmap')
        plt.title('Heatmap of Actions' if not show_q_values else 'Total Q-Values Heatmap')
        
        # Draw shortest path
        shortest_path = self.get_shortest_path()
        if shortest_path:
            path_x, path_y = zip(*shortest_path)
            plt.plot(path_y, path_x, marker='o', color='cyan', label='Shortest Path')
        
        for point in self.blocked_points:
            plt.scatter(point[1], point[0], marker='s', color='purple', label='Blocked' 
                        if point == self.blocked_points[0] else "", zorder=2)
        
        plt.scatter(self.start[1], self.start[0], marker='o', color='green', label='Start', zorder=3)
        plt.scatter(self.end[1], self.end[0], marker='x', color='blue', label='End', zorder=3)

        plt.legend()
        plt.grid(True)
        plt.xticks(np.arange(self.grid_size))
        plt.yticks(np.arange(self.grid_size))
        plt.show()
        plt.close()
        plt.savefig(f"heatmap_qvalues_{self.grid_size}.png")  # Save heatmap
        plt.close()

        # Print Q-Values table for each component
        if show_q_values:
            print("Decomposed Q-Values for each action:")
            for c_idx, component in enumerate(self.reward_components):
                print(f"Component: {component}")
                for i in range(self.grid_size):
                    for j in range(self.grid_size):
                        print(f"State {i, j}: Q-Values: {self.q_table[c_idx, i, j]}")
    def smooth(self, data, window=50):
        return np.convolve(data, np.ones(window)/window, mode='valid')


    def plot_metrics(self, save_path=None):
        episodes = list(range(1, len(self.reward_history) + 1))
        window = 1000  # larger smoothing window for smoother curve

        # Configure font
        plt.rcParams.update({'font.size': 12})

        fig, axs = plt.subplots(1, 2, figsize=(14, 5))

        # --- 1. Reward plot ---
        axs[0].plot(episodes, self.reward_history, label='Total Reward per Episode', color='cornflowerblue', alpha=0.4)
        
        if len(self.reward_history) >= window:
            smoothed = np.convolve(self.reward_history, np.ones(window)/window, mode='valid')
            axs[0].plot(episodes[len(episodes)-len(smoothed):], smoothed, label='Smoothed Reward', color='red', linewidth=2)

        axs[0].set_title('Learning Curve: Total Reward per Episode')
        axs[0].set_xlabel('Episode')
        axs[0].set_ylabel('Total Reward')
        axs[0].legend()
        axs[0].grid(True)

        # --- 2. Q-values heatmap ---
        total_q = np.sum(self.q_table, axis=0)
        q_values = np.max(total_q, axis=2)
        im = axs[1].imshow(q_values, cmap='hot', interpolation='nearest')
        axs[1].set_title('Total Q-Values Heatmap')
        fig.colorbar(im, ax=axs[1], label='Total Q-Values')

        path = self.get_shortest_path()
        if path:
            path_x, path_y = zip(*path)
            axs[1].plot(path_y, path_x, marker='o', color='cyan', label='Shortest Path')

        for point in self.blocked_points:
            axs[1].scatter(point[1], point[0], marker='s', color='purple', s=15)
        axs[1].scatter(self.start[1], self.start[0], marker='o', color='green', label='Start')
        axs[1].scatter(self.end[1], self.end[0], marker='x', color='blue', label='End')
        axs[1].legend()
        axs[1].grid(True)

        plt.tight_layout()

        if save_path:
            plt.savefig(save_path, format='pdf')  # Save for paper
            print(f"[✓] Saved plot to {save_path}")
        else:
            plt.show()
    def visualize_optimal_policy(self):
        arrow_map = {0: '↑', 1: '↓', 2: '←', 3: '→'}
        policy_grid = np.empty((self.grid_size, self.grid_size), dtype=str)

        for i in range(self.grid_size):
            for j in range(self.grid_size):
                if (i, j) in self.blocked_points:
                    policy_grid[i, j] = '█'  # blocked
                elif (i, j) == self.end:
                    policy_grid[i, j] = 'G'  # goal
                elif (i, j) == self.start:
                    policy_grid[i, j] = 'S'  # start
                else:
                    total_q = np.sum(self.q_table[:, i, j, :], axis=0)
                    best_action = np.argmax(total_q)
                    policy_grid[i, j] = arrow_map[best_action]

        print("\nOptimal Policy Grid (based on total Q-values):")
        for row in policy_grid:
            print(' '.join(row))
    def plot_policy_arrows(self):
        total_q = np.sum(self.q_table, axis=0)
        fig, ax = plt.subplots(figsize=(6, 6))
        ax.set_xlim(0, self.grid_size)
        ax.set_ylim(0, self.grid_size)
        ax.set_xticks(np.arange(0, self.grid_size + 1, 1))
        ax.set_yticks(np.arange(0, self.grid_size + 1, 1))
        ax.set_xticklabels(np.arange(0, self.grid_size + 1, 1))  # x-axis 0 → 10
        ax.set_yticklabels(np.arange(0, self.grid_size + 1, 1))  # y-axis 0 → 10 # Labels from left to right 0 → 10
        ax.grid(True)

        for i in range(self.grid_size):
            for j in range(self.grid_size):
                cell_y = self.grid_size - 1 - i  # convert coordinates so origin (0,0) is top-left

                if (i, j) in self.blocked_points:
                    ax.add_patch(plt.Rectangle((j, cell_y), 1, 1, color='black'))
                elif (i, j) == self.start:
                    ax.text(j + 0.5, cell_y + 0.5, 'S', ha='center', va='center', fontsize=12, color='green')
                elif (i, j) == self.end:
                    ax.text(j + 0.5, cell_y + 0.5, 'G', ha='center', va='center', fontsize=12, color='blue')
                else:
                    best_action = np.argmax(total_q[i, j, :])
                    dx, dy = [(0, 0.3), (0, -0.3), (-0.3, 0), (0.3, 0)][best_action]
                    ax.arrow(j + 0.5, cell_y + 0.5, dx, dy,
                            head_width=0.2, head_length=0.2, fc='red', ec='red')

        ax.set_title("Optimal Policy Arrows")
        plt.savefig(f"policy_{self.grid_size}.png", dpi=300)
        plt.show()
    def get_roulette_probabilities(self, state):
        probs_dict = {}

        for c_idx, component in enumerate(self.reward_components + ['total']):
            if component == 'total':
                q_values = np.sum(self.q_table[:, state[0], state[1], :], axis=0)
            else:
                q_values = self.q_table[c_idx, state[0], state[1], :]

            # Shift to positive to avoid negative numbers
            q_shifted = q_values - np.min(q_values) + 1e-6
            probs = q_shifted / np.sum(q_shifted)
            
            probs_dict[component] = probs  # 4 values corresponding to [up, down, left, right]

        return probs_dict

    # --- MAIN EVALUATION METRICS SECTION ---
    def _vector_from_explanation(self, explanation):
        """Return contributions vector according to reward_components order"""
        return np.array([explanation[c] for c in self.reward_components], dtype=float)

    def evaluate_explanations(self, Seval=None, save_path=None, suff_threshold=0.9):
        """
        Evaluate explanations with main metrics:
        - Action Agreement (AA)    <- using E_all (msx+ ∪ msx-)
        - Sufficiency (as continuous score + binary by threshold) <- NOW using E_all (msx+ ∪ msx-)
        - Comprehensiveness <- removing E_all
        - Average Explanation Size (Compactness)
        """
        if Seval is None:
            Seval = [(i,j) for i in range(self.grid_size) for j in range(self.grid_size) 
                    if (i,j) not in self.blocked_points]
        
        per_state_metrics = []
        all_action_agreement = []
        all_sufficiency = []
        all_suff_binary = []
        all_comprehensiveness = []
        all_aes = []
        
        for state in Seval:
            row = {"State": str(state)}
            total_q = np.sum(self.q_table[:, state[0], state[1], :], axis=0)
            chosen_action = int(np.argmax(total_q))
            chosen_action_q = float(total_q[chosen_action])
            
            available = self.get_actions(state)
            if len(available) > 1:
                alt = random.choice([a for a in available if a != chosen_action])
            else:
                alt = chosen_action
            
            # Get MSX explanation
            msx_plus, msx_minus, explanation, _ = self.minimal_sufficient_explanation(state, chosen_action, alt)
            E_plus = set(msx_plus)                # positive explanatory components
            E_minus = set(msx_minus)              # negative explanatory components
            E_all = E_plus.union(E_minus)         # <-- now Sufficiency uses E_all as required
            msx_size = len(msx_plus) + len(msx_minus)
            all_aes.append(msx_size)
            row["MSX_size"] = msx_size
            row["MSX_plus"] = str(msx_plus)
            row["MSX_minus"] = str(msx_minus)

            # 1. ACTION AGREEMENT - surrogate policy using E_all (both msx+ and msx-)
            q_e = np.zeros(4)
            for action in range(4):
                for c_idx, comp in enumerate(self.reward_components):
                    if comp in E_all:
                        q_e[action] += self.q_table[c_idx, state[0], state[1], action]
            # add very small noise to break deterministic ties
            q_e = q_e + np.random.normal(scale=1e-9, size=q_e.shape)
            action_e = int(np.argmax(q_e))
            aa_score = 1 if action_e == chosen_action else 0
            all_action_agreement.append(aa_score)
            row["Action_Agreement"] = aa_score

            # 2. SUFFICIENCY - NOW measure fraction of chosen_action's Q retained when only keeping E_all (msx+ U msx-)
            total_q_keepE_all = np.zeros_like(total_q)
            for a in range(4):
                ssum = 0.0
                for c_idx, comp in enumerate(self.reward_components):
                    if comp in E_all:
                        ssum += self.q_table[c_idx, state[0], state[1], a]
                total_q_keepE_all[a] = ssum

            # sufficiency score for chosen action (using E_all)
            denom = chosen_action_q if abs(chosen_action_q) > 1e-9 else 1.0
            suff_score = float(total_q_keepE_all[chosen_action]) / denom
            # clamp to avoid infinite or nan values
            if math.isnan(suff_score) or math.isinf(suff_score):
                suff_score = 0.0
            # binary by threshold
            suff_binary = 1 if suff_score >= suff_threshold else 0

            row["Sufficiency"] = suff_score   # save score (continuous)
            row["Sufficiency_binary"] = suff_binary
            all_sufficiency.append(suff_score)
            all_suff_binary.append(suff_binary)

            # 3. COMPREHENSIVENESS - measure drop when REMOVING E_all (both msx+ and msx-)
            total_q_excluded = np.zeros_like(total_q)
            for a in range(4):
                ssum = 0.0
                for c_idx, comp in enumerate(self.reward_components):
                    if comp not in E_all:  # remove E_all -> keep remaining part
                        ssum += self.q_table[c_idx, state[0], state[1], a]
                total_q_excluded[a] = ssum
            
            comp_drop = chosen_action_q - total_q_excluded[chosen_action]
            # Normalize
            total_q_grid = np.sum(self.q_table, axis=0)
            max_q = np.max(total_q_grid)
            min_q = np.min(total_q_grid)
            denom2 = (max_q - min_q) if (max_q - min_q) != 0 else 1.0
            comprehensiveness_norm = comp_drop / denom2
            row["Comprehensiveness"] = comprehensiveness_norm
            all_comprehensiveness.append(comprehensiveness_norm)

            # Save explanation values (for debug)
            for c_idx, comp in enumerate(self.reward_components):
                row[f"Δ_{comp}"] = float(explanation[comp])

            per_state_metrics.append(row)

        # Summary metrics
        summary = {
            "Action_Agreement": float(np.mean(all_action_agreement)) if all_action_agreement else 0.0,
            "Sufficiency_score_mean": float(np.mean(all_sufficiency)) if all_sufficiency else 0.0,
            "Sufficiency_binary_mean": float(np.mean(all_suff_binary)) if all_suff_binary else 0.0,
            "Comprehensiveness": float(np.mean(all_comprehensiveness)) if all_comprehensiveness else 0.0,
            "AES": float(np.mean(all_aes)) if all_aes else 0.0,
            "NumStatesEvaluated": len(Seval)
        }

        # Save results
        if save_path is None:
            run_count = self.get_run_count(self.grid_size)
            save_path = self.prepare_output_file(self.grid_size, run_count)

        df_per_state = pd.DataFrame(per_state_metrics)
        
        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            df_per_state.to_excel(writer, sheet_name='PerState_Metrics', index=False)
            pd.DataFrame([summary]).to_excel(writer, sheet_name='Summary', index=False)

        print(f"[✓] Evaluation saved to {save_path}")
        print("Summary Metrics:")
        for key, value in summary.items():
            if isinstance(value, float):
                print(f"  {key}: {value:.4f}")
            else:
                print(f"  {key}: {value}")

        return summary, df_per_state

def plot_reward_curve_drq_rws(file_path, window=100):
    df = pd.read_excel(file_path, sheet_name='Episode Log')
    
    rewards = []
    episode_reward = 0

    for _, row in df.iterrows():
        state = ast.literal_eval(row["State"])  # Convert string to tuple
        reward = row["Total Δ"]
        if state == (0, 0) and episode_reward > 0:
            rewards.append(episode_reward)
            episode_reward = 0
        episode_reward += reward
    if episode_reward > 0:
        rewards.append(episode_reward)

    episodes = list(range(1, len(rewards) + 1))

    def smooth(data, w):
        return np.convolve(data, np.ones(w)/w, mode='valid')

    plt.figure(figsize=(10, 5))
    plt.plot(episodes, rewards, label="DRQ + RWS (raw)", alpha=0.4)
    if len(rewards) >= window:
        smoothed = smooth(rewards, window)
        plt.plot(episodes[window-1:], smoothed, label="DRQ + RWS (smoothed)", linewidth=2)

    plt.title("Reward per Episode - DRQ + RWS (Fixed)")
    plt.xlabel("Episode")
    plt.ylabel("Total Reward")
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.savefig("reward_curve_DRQ_RWS_FIXED.png")
    plt.show()
def is_path_available(start, end, blocked, grid_size):
            visited = set()
            queue = deque([start])
            while queue:
                current = queue.popleft()
                if current == end:
                    return True
                for dx, dy in [(-1,0),(1,0),(0,-1),(0,1)]:
                    nx, ny = current[0]+dx, current[1]+dy
                    if 0 <= nx < grid_size and 0 <= ny < grid_size:
                        if (nx, ny) not in blocked and (nx, ny) not in visited:
                            visited.add((nx, ny))
                            queue.append((nx, ny))
            return False

#State classification and group analysis
def analyze_and_group_states(excel_file: str, maze_file: str, grid_size: int, save_path: str = None):
    """
    Group states and analyze FULL metrics by group
    Each group has complete metrics similar to Summary
    """
    import pandas as pd
    import numpy as np
    import matplotlib.pyplot as plt
    import os

    def _load_blocked_from_mazefile(path):
        blocked = []
        with open(path, 'r') as f:
            for i, line in enumerate(f):
                for j, ch in enumerate(line.strip().split()):
                    if ch == '#':
                        blocked.append((i, j))
        return blocked

    def _manhattan(a, b):
        return abs(a[0]-b[0]) + abs(a[1]-b[1])

    def _neighbors(state, grid_size):
        i, j = state
        res = []
        for di, dj in [(-1,0),(1,0),(0,-1),(0,1)]:
            ni, nj = i+di, j+dj
            if 0 <= ni < grid_size and 0 <= nj < grid_size:
                res.append((ni, nj))
        return res

    # Load data
    xls = pd.ExcelFile(excel_file)
    if 'PerState_Metrics' not in xls.sheet_names:
        raise ValueError("Excel file must contain 'PerState_Metrics' sheet")
    
    per_state = pd.read_excel(xls, 'PerState_Metrics')
    summary_df = pd.read_excel(xls, 'Summary') if 'Summary' in xls.sheet_names else None
    
    blocked = set(_load_blocked_from_mazefile(maze_file))
    end = (grid_size - 1, grid_size - 1)

    per_state = per_state.copy()
    per_state['state_tup'] = per_state['State'].apply(lambda s: eval(s) if isinstance(s, str) else tuple(s))

    all_states = [(i, j) for i in range(grid_size) for j in range(grid_size) if (i, j) not in blocked]
    rows = []
    
    for s in all_states:
        row = {'state': s}
        i, j = s
        
        # Group classification
        row['near_wall'] = int(i == 0 or i == grid_size-1 or j == 0 or j == grid_size-1)
        row['near_obstacle'] = int(any(nb in blocked for nb in _neighbors(s, grid_size)))
        row['near_goal'] = int(_manhattan(s, end) <= 2)
        avail = [nb for nb in _neighbors(s, grid_size) if nb not in blocked]
        row['available_actions'] = len(avail)
        row['dead_end'] = int(len(avail) == 1)
        row['crossroads'] = int(len(avail) >= 3)

        # Get FULL metrics from per_state
        try:
            pr = per_state[per_state['state_tup'] == s]
            if len(pr) > 0:
                pr = pr.iloc[0]
                # Get all main metrics similar to summary
                metrics_to_extract = ['Action_Agreement', 'Sufficiency', 'Comprehensiveness', 'MSX_size']
                for col in metrics_to_extract:
                    if col in pr.index:
                        row[col] = pr[col]
                    else:
                        row[col] = np.nan
            else:
                for col in ['Action_Agreement', 'Sufficiency', 'Comprehensiveness', 'MSX_size']:
                    row[col] = np.nan
        except Exception:
            for col in ['Action_Agreement', 'Sufficiency', 'Comprehensiveness', 'MSX_size']:
                row[col] = np.nan

        rows.append(row)

    df_all = pd.DataFrame(rows)

    # Define groups
    groups = {
        'All': df_all.index >= 0,
        'NearWall': df_all['near_wall'] == 1,
        'NearObstacle': df_all['near_obstacle'] == 1,
        'NearGoal': df_all['near_goal'] == 1,
        'DeadEnd': df_all['dead_end'] == 1,
        'Crossroads': df_all['crossroads'] == 1
    }

  
    summary_rows = []
    
    for gname, mask in groups.items():
        sel = df_all[mask]
        summary = {
            'Group': gname, 
            'Count': int(len(sel)),
            'Action_Agreement': float(sel['Action_Agreement'].mean()) if len(sel) > 0 else 0,
            'Sufficiency': float(sel['Sufficiency'].mean()) if len(sel) > 0 else 0,
            'Comprehensiveness': float(sel['Comprehensiveness'].mean()) if len(sel) > 0 else 0,
            'AES': float(sel['MSX_size'].mean()) if len(sel) > 0 else 0
        }
        
 
        summary['Action_Agreement_std'] = float(sel['Action_Agreement'].std()) if len(sel) > 0 else 0
        summary['Sufficiency_std'] = float(sel['Sufficiency'].std()) if len(sel) > 0 else 0
        summary['Comprehensiveness_std'] = float(sel['Comprehensiveness'].std()) if len(sel) > 0 else 0
        summary['AES_std'] = float(sel['MSX_size'].std()) if len(sel) > 0 else 0
        
        summary_rows.append(summary)

    df_groups = pd.DataFrame(summary_rows)

    # Save results
    save_to = save_path if save_path is not None else excel_file

    try:
        from openpyxl import load_workbook
        if os.path.exists(save_to):
            wb = load_workbook(save_to)
            for sheet in ['Group_Summary', 'PerState_WithGroups']:
                if sheet in wb.sheetnames:
                    wb.remove(wb[sheet])
            wb.save(save_to)
    except Exception:
        pass

    mode = 'a' if os.path.exists(save_to) else 'w'
    with pd.ExcelWriter(save_to, engine='openpyxl', mode=mode) as writer:
        df_groups.to_excel(writer, sheet_name='Group_Summary', index=False)
        df_all.to_excel(writer, sheet_name='PerState_WithGroups', index=False)

    outdir = os.path.dirname(save_to) or '.'
    
    try:
        # Plot comparing all metrics by group
        metrics_to_plot = ['Action_Agreement', 'Sufficiency', 'Comprehensiveness', 'AES']
        groups_to_plot = [g for g in groups.keys() if g != 'All']
        
        fig, axes = plt.subplots(2, 2, figsize=(15, 10))
        axes = axes.ravel()
        
        for idx, metric in enumerate(metrics_to_plot):
            values = [df_groups[df_groups['Group'] == g][metric].values[0] for g in groups_to_plot]
            axes[idx].bar(groups_to_plot, values)
            axes[idx].set_title(f'{metric} by Group')
            axes[idx].set_ylabel(metric)
            axes[idx].tick_params(axis='x', rotation=45)
        
        plt.tight_layout()
        plt.savefig(os.path.join(outdir, f'group_comparison_grid{grid_size}.png'))
        plt.close()
        
        # Radar chart for overall comparison
        fig = plt.figure(figsize=(10, 8))
        ax = fig.add_subplot(111, polar=True)
        
        # Normalize metrics to 0-1 scale
        normalized_metrics = {}
        for metric in metrics_to_plot:
            max_val = df_groups[metric].max()
            min_val = df_groups[metric].min()
            if max_val != min_val:
                normalized_metrics[metric] = [(df_groups[df_groups['Group'] == g][metric].values[0] - min_val) / (max_val - min_val) 
                                           for g in groups_to_plot]
            else:
                normalized_metrics[metric] = [0.5 for _ in groups_to_plot]
        
        # Draw radar chart
        angles = np.linspace(0, 2*np.pi, len(metrics_to_plot), endpoint=False).tolist()
        angles += angles[:1]  # Close shape
        
        for i, group in enumerate(groups_to_plot):
            values = [normalized_metrics[metric][i] for metric in metrics_to_plot]
            values += values[:1]  # Close shape
            ax.plot(angles, values, 'o-', linewidth=2, label=group)
            ax.fill(angles, values, alpha=0.1)
        
        ax.set_xticks(angles[:-1])
        ax.set_xticklabels(metrics_to_plot)
        ax.set_ylim(0, 1)
        ax.legend(bbox_to_anchor=(1.1, 1.1))
        plt.title('Normalized Metrics Comparison by Group (Radar Chart)')
        plt.tight_layout()
        plt.savefig(os.path.join(outdir, f'group_radar_grid{grid_size}.png'))
        plt.close()
        
    except Exception as e:
        print(f"Could not create comparison plots: {e}")

    print(f"[✓] Group analysis with full metrics saved to {save_to}")
    return df_groups, df_all


def print_summary(agent, label, train_time, selection_mode="Roulette Wheel Selection"):
    total_episodes = len(agent.reward_history)
    avg_reward_all = np.mean(agent.reward_history)
    avg_reward_last_100 = np.mean(agent.reward_history[-100:]) if total_episodes >= 100 else avg_reward_all
    turn_ratio = agent.final_turns / agent.final_steps if agent.final_steps > 0 else 0.0

    print(f"\n=== Summary for {label} ===")
    print(f"    Selection Mode           : {selection_mode}")
    print(f"    Final Reward             : {agent.final_reward}")
    print(f"    Avg Reward (all)         : {avg_reward_all:.2f}")
    print(f"    Avg Reward (last 100)    : {avg_reward_last_100:.2f}")
    print(f"    Total Steps to Goal      : {agent.final_steps}")
    print(f"    Total Turns              : {agent.final_turns}")
    print(f"    Turn Ratio (180° / step) : {turn_ratio:.3f}")
    print(f"    Training Episodes        : {total_episodes}")
    print(f"    Training Time (sec)      : {train_time:.2f}")
    print("-" * 60)




if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python drQ-main.py <maze_file>")
        sys.exit(1)

    maze_file = sys.argv[1]
    grid_size = 10
    agent = RewardPathFinder(grid_size, maze_file)
    run_count = agent.get_run_count(grid_size)

    # Create output file name
    output_file = "excel-results/evaluation-report-final.xlsx"

    # Train agent
    start_time = time.time()
    agent.train(3000,log_random_episode=True)

    # --- ADD EXPLANATION EVALUATION AND GROUPING ---
    print("\n=== Evaluating Explanations ===")
    summary, per_state_df = agent.evaluate_explanations(save_path=output_file)
    print("Evaluation summary:")
    print(summary)

    print("\n=== Grouping States ===")
    group_df, perstate_with_groups = analyze_and_group_states(output_file, maze_file, grid_size, save_path=None)
    print("Group-level summary (first rows):")
    print(group_df.head())

    # Display results and save log
    agent.visualize(show_q_values=True)
    agent.plot_policy_arrows()
    print(f"Output file: {output_file}")
    end_time = time.time()
    elapsed = end_time - start_time
    print(f"Training time: {elapsed:.2f} seconds")
    plot_reward_curve_drq_rws("excel-results/grid-20-log-13.xlsx")
    print_summary(agent, label="DRQ + RWS", train_time=elapsed)

# python drQ-main.py maze10.txt
# python drQ-main.py maze20.txt
# python drQ-main.py maze30.txt